# This will be able to automatically process thousands of sheets in under a second
# uses openpyxl

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    # renaming the function as wb to make it easier
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # for loop to generate the cell numbers, then add the new cell into a different column.
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
